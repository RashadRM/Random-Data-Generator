import random
import string
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import json
from openpyxl import Workbook
import numpy as np


class RandomDataGenerator:
    def __init__(self):
        self.columns = []
        self.data = []

    def add_column(self, name, data_type, value_type, **kwargs):
        column = {'name': name, 'data_type': data_type, 'value_type': value_type, 'kwargs': kwargs}
        self.columns.append(column)

    def generate_data(self, num_rows):
        self.data = []
        for _ in range(num_rows):
            row = {}
            for column in self.columns:
                name = column['name']
                data_type = column['data_type']
                value_type = column['value_type']
                kwargs = column['kwargs']

                if data_type == 'string':
                    length = kwargs.get('length', 10)
                    include_numbers = kwargs.get('include_numbers', False)
                    include_special = kwargs.get('include_special', False)
                    characters = string.ascii_letters
                    if include_numbers:
                        characters += string.digits
                    if include_special:
                        characters += string.punctuation
                    value = ''.join(random.choices(characters, k=length))
                elif data_type == 'numeric':
                    mean = kwargs.get('mean', 0)
                    stddev = kwargs.get('stddev', 1)
                    distribution = kwargs.get('distribution', 'uniform')
                    is_float = kwargs.get('is_float', False)
                    if distribution == 'normal':
                        value = np.random.normal(mean, stddev)
                    else:
                        start = kwargs.get('start', 0)
                        end = kwargs.get('end', 100)
                        if is_float:
                            value = random.uniform(start, end)
                        else:
                            value = random.randint(start, end)
                elif data_type == 'date':
                    start_date = kwargs.get('start_date', '1970-01-01')
                    end_date = kwargs.get('end_date', 'now')
                    format = kwargs.get('format', '%Y-%m-%d')
                    start_date = datetime.strptime(start_date, format)
                    if end_date == 'now':
                        end_date = datetime.now()
                    else:
                        end_date = datetime.strptime(end_date, format)
                    value = start_date + timedelta(
                        seconds=random.randint(0, int((end_date - start_date).total_seconds())))
                    value = value.strftime(format)
                elif data_type == 'custom':
                    custom_values = kwargs.get('custom_values', [])
                    value = random.choice(custom_values)
                elif data_type == 'id':
                    value = _  # Unique ID is just the row index
                elif data_type == 'phone':
                    country_code = kwargs.get('country_code', '+1')
                    number_length = kwargs.get('number_length', 10)
                    value = country_code + ''.join(random.choices(string.digits, k=number_length))
                elif data_type == 'country':
                    value = random.choice(
                        ['USA', 'Canada', 'UK', 'Germany', 'France', 'Australia', 'India', 'China', 'Japan', 'Brazil'])

                row[name] = value
            self.data.append(row)

        return self.data

    def save_to_file(self, filename, file_type):
        if file_type == 'csv':
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=[col['name'] for col in self.columns])
                writer.writeheader()
                for row in self.data:
                    writer.writerow(row)
        elif file_type == 'xlsx':
            wb = Workbook()
            ws = wb.active
            for col_index, column in enumerate(self.columns, start=1):
                ws.cell(row=1, column=col_index, value=column['name'])
                for row_index, row in enumerate(self.data, start=2):
                    ws.cell(row=row_index, column=col_index, value=row[column['name']])
            wb.save(filename)
        elif file_type == 'txt':
            with open(filename, 'w') as txtfile:
                for row in self.data:
                    txtfile.write(str(row) + '\n')
        elif file_type == 'json':
            with open(filename, 'w') as jsonfile:
                json.dump(self.data, jsonfile)
        else:
            # Handle unsupported file types
            pass


class RandomDataGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Random Data Generator")

        self.generator = RandomDataGenerator()

        self.frame = ttk.Frame(root, padding="20")
        self.frame.grid(row=0, column=0, sticky="nsew")

        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.frame, text="Column Name:").grid(row=0, column=0, padx=5, pady=5)
        self.column_name_entry = ttk.Entry(self.frame)
        self.column_name_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self.frame, text="Data Type:").grid(row=1, column=0, padx=5, pady=5)
        self.data_type_combo = ttk.Combobox(self.frame,
                                            values=["string", "numeric", "date", "custom", "id", "phone", "country"])
        self.data_type_combo.grid(row=1, column=1, padx=5, pady=5)
        self.data_type_combo.bind("<<ComboboxSelected>>", self.on_data_type_selected)

        ttk.Label(self.frame, text="Value Type:").grid(row=2, column=0, padx=5, pady=5)
        self.value_type_combo = ttk.Combobox(self.frame, values=["random", "sequential"])
        self.value_type_combo.grid(row=2, column=1, padx=5, pady=5)

        self.random_options_frame = ttk.Frame(self.frame)
        self.random_options_frame.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="w")

        self.start_label = ttk.Label(self.random_options_frame, text="Start:")
        self.start_label.grid(row=0, column=0, padx=5, pady=5)
        self.start_entry = ttk.Entry(self.random_options_frame, width=12)
        self.start_entry.grid(row=0, column=1, padx=5, pady=5)

        self.end_label = ttk.Label(self.random_options_frame, text="End:")
        self.end_label.grid(row=1, column=0, padx=5, pady=5)
        self.end_entry = ttk.Entry(self.random_options_frame, width=12)
        self.end_entry.grid(row=1, column=1, padx=5, pady=5)

        self.format_label = ttk.Label(self.random_options_frame, text="Date Format:")
        self.format_label.grid(row=2, column=0, padx=5, pady=5)
        self.format_combo = ttk.Combobox(self.random_options_frame, values=["%Y-%m-%d", "%m-%d-%Y", "%d-%m-%Y"])
        self.format_combo.grid(row=2, column=1, padx=5, pady=5)
        self.format_combo.current(0)

        self.length_label = ttk.Label(self.random_options_frame, text="Length:")
        self.length_label.grid(row=3, column=0, padx=5, pady=5)
        self.length_entry = ttk.Entry(self.random_options_frame)
        self.length_entry.grid(row=3, column=1, padx=5, pady=5)

        self.custom_values_label = ttk.Label(self.random_options_frame, text="Custom Values (comma-separated):")
        self.custom_values_label.grid(row=4, column=0, padx=5, pady=5)
        self.custom_values_entry = ttk.Entry(self.random_options_frame, width=25)
        self.custom_values_entry.grid(row=4, column=1, padx=5, pady=5)

        self.include_numbers_var = tk.BooleanVar()
        self.include_numbers_check = ttk.Checkbutton(self.random_options_frame, text="Include Numbers",
                                                     variable=self.include_numbers_var)
        self.include_numbers_check.grid(row=5, column=0, padx=5, pady=5, sticky="w")

        self.include_special_var = tk.BooleanVar()
        self.include_special_check = ttk.Checkbutton(self.random_options_frame, text="Include Special Characters",
                                                     variable=self.include_special_var)
        self.include_special_check.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        self.mean_label = ttk.Label(self.random_options_frame, text="Mean:")
        self.mean_label.grid(row=6, column=0, padx=5, pady=5)
        self.mean_entry = ttk.Entry(self.random_options_frame)
        self.mean_entry.grid(row=6, column=1, padx=5, pady=5)

        self.stddev_label = ttk.Label(self.random_options_frame, text="StdDev:")
        self.stddev_label.grid(row=7, column=0, padx=5, pady=5)
        self.stddev_entry = ttk.Entry(self.random_options_frame)
        self.stddev_entry.grid(row=7, column=1, padx=5, pady=5)

        self.distribution_label = ttk.Label(self.random_options_frame, text="Distribution:")
        self.distribution_label.grid(row=8, column=0, padx=5, pady=5)
        self.distribution_combo = ttk.Combobox(self.random_options_frame, values=["uniform", "normal"])
        self.distribution_combo.grid(row=8, column=1, padx=5, pady=5)
        self.distribution_combo.current(0)

        self.add_column_button = ttk.Button(self.frame, text="Add Column", command=self.add_column)
        self.add_column_button.grid(row=9, column=0, padx=5, pady=5)

        self.added_column_label = ttk.Label(self.frame, text="")
        self.added_column_label.grid(row=10, column=0, columnspan=2, padx=5, pady=5)

        ttk.Label(self.frame, text="Number of Rows:").grid(row=11, column=0, padx=5, pady=5)
        self.num_rows_entry = ttk.Entry(self.frame)
        self.num_rows_entry.grid(row=11, column=1, padx=5, pady=5)

        self.preview_button = ttk.Button(self.frame, text="Preview Data", command=self.preview_data)
        self.preview_button.grid(row=12, column=0, padx=5, pady=5)

        self.generate_button = ttk.Button(self.frame, text="Generate Data", command=self.generate_data)
        self.generate_button.grid(row=12, column=1, padx=5, pady=5)

        ttk.Label(self.frame, text="Save As:").grid(row=13, column=0, padx=5, pady=5)
        self.save_type_combo = ttk.Combobox(self.frame, values=["csv", "xlsx", "txt", "json"])
        self.save_type_combo.grid(row=13, column=1, padx=5, pady=5)
        self.save_type_combo.current(0)

        self.save_button = ttk.Button(self.frame, text="Save Data", command=self.save_data)
        self.save_button.grid(row=14, column=0, columnspan=2, padx=5, pady=5)

        self.output_text = tk.Text(self.frame, height=10, width=50)
        self.output_text.grid(row=15, column=0, columnspan=2, padx=5, pady=5)

    def on_data_type_selected(self, event):
        data_type = self.data_type_combo.get()
        self.reset_random_options()
        if data_type == "string":
            self.random_options_frame.grid()
            self.length_label.grid()
            self.length_entry.grid()
            self.include_numbers_check.grid()
            self.include_special_check.grid()
            self.start_label.grid_remove()
            self.start_entry.grid_remove()
            self.end_label.grid_remove()
            self.end_entry.grid_remove()
            self.format_label.grid_remove()
            self.format_combo.grid_remove()
            self.mean_label.grid_remove()
            self.mean_entry.grid_remove()
            self.stddev_label.grid_remove()
            self.stddev_entry.grid_remove()
            self.distribution_label.grid_remove()
            self.distribution_combo.grid_remove()
            self.custom_values_label.grid_remove()
            self.custom_values_entry.grid_remove()
        elif data_type == "numeric":
            self.random_options_frame.grid()
            self.length_label.grid_remove()
            self.length_entry.grid_remove()
            self.include_numbers_check.grid_remove()
            self.include_special_check.grid_remove()
            self.start_label.grid()
            self.start_entry.grid()
            self.end_label.grid()
            self.end_entry.grid()
            self.mean_label.grid()
            self.mean_entry.grid()
            self.stddev_label.grid()
            self.stddev_entry.grid()
            self.distribution_label.grid()
            self.distribution_combo.grid()
            self.format_label.grid_remove()
            self.format_combo.grid_remove()
            self.custom_values_label.grid_remove()
            self.custom_values_entry.grid_remove()
        elif data_type == "date":
            self.random_options_frame.grid()
            self.length_label.grid_remove()
            self.length_entry.grid_remove()
            self.include_numbers_check.grid_remove()
            self.include_special_check.grid_remove()
            self.start_label.grid()
            self.start_entry.grid()
            self.end_label.grid()
            self.end_entry.grid()
            self.format_label.grid()
            self.format_combo.grid()
            self.mean_label.grid_remove()
            self.mean_entry.grid_remove()
            self.stddev_label.grid_remove()
            self.stddev_entry.grid_remove()
            self.distribution_label.grid_remove()
            self.distribution_combo.grid_remove()
            self.custom_values_label.grid_remove()
            self.custom_values_entry.grid_remove()
        elif data_type == "custom":
            self.random_options_frame.grid()
            self.length_label.grid_remove()
            self.length_entry.grid_remove()
            self.include_numbers_check.grid_remove()
            self.include_special_check.grid_remove()
            self.start_label.grid_remove()
            self.start_entry.grid_remove()
            self.end_label.grid_remove()
            self.end_entry.grid_remove()
            self.format_label.grid_remove()
            self.format_combo.grid_remove()
            self.mean_label.grid_remove()
            self.mean_entry.grid_remove()
            self.stddev_label.grid_remove()
            self.stddev_entry.grid_remove()
            self.distribution_label.grid_remove()
            self.distribution_combo.grid_remove()
            self.custom_values_label.grid()
            self.custom_values_entry.grid()
        elif data_type == "id":
            self.random_options_frame.grid_remove()
        elif data_type == "phone":
            self.random_options_frame.grid()
            self.start_label.config(text="Country Code:")
            self.start_entry.grid()
            self.start_entry.delete(0, tk.END)
            self.start_entry.insert(0, "+1")
            self.end_label.config(text="Number Length:")
            self.end_entry.grid()
            self.end_entry.delete(0, tk.END)
            self.end_entry.insert(0, "10")
            self.length_label.grid_remove()
            self.length_entry.grid_remove()
            self.include_numbers_check.grid_remove()
            self.include_special_check.grid_remove()
            self.format_label.grid_remove()
            self.format_combo.grid_remove()
            self.mean_label.grid_remove()
            self.mean_entry.grid_remove()
            self.stddev_label.grid_remove()
            self.stddev_entry.grid_remove()
            self.distribution_label.grid_remove()
            self.distribution_combo.grid_remove()
            self.custom_values_label.grid_remove()
            self.custom_values_entry.grid_remove()
        elif data_type == "country":
            self.random_options_frame.grid_remove()

    def reset_random_options(self):
        self.start_label.grid_remove()
        self.start_entry.grid_remove()
        self.end_label.grid_remove()
        self.end_entry.grid_remove()
        self.format_label.grid_remove()
        self.format_combo.grid_remove()
        self.length_label.grid_remove()
        self.length_entry.grid_remove()
        self.custom_values_label.grid_remove()
        self.custom_values_entry.grid_remove()
        self.include_numbers_check.grid_remove()
        self.include_special_check.grid_remove()
        self.mean_label.grid_remove()
        self.mean_entry.grid_remove()
        self.stddev_label.grid_remove()
        self.stddev_entry.grid_remove()
        self.distribution_label.grid_remove()
        self.distribution_combo.grid_remove()

    def add_column(self):
        column_name = self.column_name_entry.get()
        data_type = self.data_type_combo.get()
        value_type = self.value_type_combo.get()
        kwargs = {}
        if data_type == "numeric":
            kwargs["start"] = float(self.start_entry.get())
            kwargs["end"] = float(self.end_entry.get())
            kwargs["is_float"] = (value_type == "float")
            kwargs["mean"] = float(self.mean_entry.get()) if self.mean_entry.get() else 0
            kwargs["stddev"] = float(self.stddev_entry.get()) if self.stddev_entry.get() else 1
            kwargs["distribution"] = self.distribution_combo.get()
        elif data_type == "string":
            kwargs["length"] = int(self.length_entry.get())
            kwargs["include_numbers"] = self.include_numbers_var.get()
            kwargs["include_special"] = self.include_special_var.get()
        elif data_type == "date":
            kwargs["start_date"] = self.start_entry.get()
            kwargs["end_date"] = self.end_entry.get()
            kwargs["format"] = self.format_combo.get()
        elif data_type == "custom":
            kwargs["custom_values"] = [val.strip() for val in self.custom_values_entry.get().split(",")]

        self.generator.add_column(column_name, data_type, value_type, **kwargs)

        self.column_name_entry.delete(0, tk.END)
        self.start_entry.delete(0, tk.END)
        self.end_entry.delete(0, tk.END)
        self.length_entry.delete(0, tk.END)
        self.custom_values_entry.delete(0, tk.END)

        self.added_column_label.config(text=f"Column '{column_name}' added.")

    def generate_data(self):
        self.output_text.delete('1.0', tk.END)
        num_rows = int(self.num_rows_entry.get())
        data = self.generator.generate_data(num_rows)
        for row in data:
            self.output_text.insert(tk.END, str(row) + "\n")

    def preview_data(self):
        try:
            num_rows = int(self.num_rows_entry.get())
        except ValueError:
            tk.messagebox.showerror("Error", "Please enter a valid number of rows.")
            return

        if num_rows <= 0:
            tk.messagebox.showerror("Error", "Number of rows must be greater than 0.")
            return

        data = self.generator.generate_data(num_rows)

        # Create a new window for preview data
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Preview Data")

        # Add a table to the new window
        columns = [col['name'] for col in self.generator.columns]
        tree = ttk.Treeview(preview_window, columns=columns, show='headings')
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, minwidth=0, width=100, stretch=tk.NO)

        for row in data:
            tree.insert("", tk.END, values=[row[col] for col in columns])

        tree.pack(expand=True, fill=tk.BOTH)

        # Add a scrollbar to the table
        scrollbar = ttk.Scrollbar(preview_window, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def save_data(self):
        file_type = self.save_type_combo.get()
        filename = filedialog.asksaveasfilename(defaultextension=f".{file_type}",
                                                filetypes=[(file_type.upper(), f"*.{file_type}")])
        if filename:
            self.generator.save_to_file(filename, file_type)
            messagebox.showinfo("Save Data", f"Data saved successfully as {filename}")


if __name__ == "__main__":
    root = tk.Tk()
    app = RandomDataGeneratorApp(root)
    root.mainloop()
